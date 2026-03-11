"""
Excel 결과 출력 모듈
- Sheet1: 파일별 상세 데이터
- Sheet2: 폴더별 통계 (8회 평균)
"""
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.formatting.rule import ColorScaleRule
import config


# ── 스타일 상수 ─────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="5B21B6")   # 진보라
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SUBHEADER_FILL = PatternFill("solid", fgColor="EDE9FE") # 연보라
SUBHEADER_FONT = Font(bold=True, color="4C1D95", size=10)
EXCLUDED_FILL = PatternFill("solid", fgColor="FEF3C7")  # 연노랑 (제외 행)
EXCLUDED_FONT = Font(color="92400E", italic=True)
ERROR_FILL = PatternFill("solid", fgColor="FEE2E2")     # 연빨강 (오류 행)
BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def export(folder_results: dict, folder_stats: dict, output_path: str = None) -> str:
    """
    분석 결과를 Excel 파일로 저장합니다.

    Args:
        folder_results: {명령어: [AnalysisResult, ...]}
        folder_stats: {명령어: FolderStats}
        output_path: 저장 경로 (None이면 자동 생성)

    Returns:
        저장된 파일 경로
    """
    if output_path is None:
        os.makedirs(config.EXCEL_DIR, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(config.EXCEL_DIR, f"E2E_분석결과_{timestamp}.xlsx")

    wb = openpyxl.Workbook()

    # ── Sheet1: 파일별 상세 ──────────────────────────────
    ws1 = wb.active
    ws1.title = "파일별 상세"
    _write_detail_sheet(ws1, folder_results, folder_stats)

    # ── Sheet2: 폴더별 통계 ─────────────────────────────
    ws2 = wb.create_sheet("폴더별 통계")
    _write_stats_sheet(ws2, folder_stats)

    wb.save(output_path)
    return output_path


def _write_detail_sheet(ws, folder_results: dict, folder_stats: dict):
    """Sheet1: 파일별 T0~T3, E2E 상세 데이터"""
    headers = ["명령어", "파일명", "T0 (s)", "T1 (s)", "T2 (s)", "T3 (s)", "최종응답속도 E2E (s)", "비고"]
    col_widths = [28, 30, 10, 10, 10, 10, 20, 15]

    _write_header_row(ws, 1, headers, HEADER_FILL, HEADER_FONT)
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    row = 2
    for command, results in folder_results.items():
        stats = folder_stats.get(command)
        excluded_set = set(stats.excluded_indices) if stats else set()

        for idx, r in enumerate(results):
            note = ""
            fill = None
            font = None

            if r.error:
                note = f"오류: {r.error[:30]}"
                fill = ERROR_FILL
                font = Font(color="991B1B")
            elif idx in excluded_set:
                # 최대/최소 제외 여부 확인
                e2e_vals = [x.E2E for x in results if x.error is None]
                if e2e_vals:
                    if r.E2E == max(e2e_vals):
                        note = "제외(최대)"
                    elif r.E2E == min(e2e_vals):
                        note = "제외(최소)"
                    else:
                        note = "제외"
                fill = EXCLUDED_FILL
                font = EXCLUDED_FONT

            values = [
                command,
                r.file_name,
                r.T0 if not r.error else "-",
                r.T1 if not r.error else "-",
                r.T2 if not r.error else "-",
                r.T3 if not r.error else "-",
                r.E2E if not r.error else "-",
                note,
            ]
            _write_data_row(ws, row, values, fill=fill, font=font)
            row += 1

        # 폴더 구분선
        row += 1

    # E2E 컬럼 조건부 서식 (컬러스케일)
    e2e_col = "G"
    if row > 2:
        ws.conditional_formatting.add(
            f"{e2e_col}2:{e2e_col}{row}",
            ColorScaleRule(
                start_type="min", start_color="63BE7B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="F8696B",
            )
        )

    ws.freeze_panes = "A2"


def _write_stats_sheet(ws, folder_stats: dict):
    """Sheet2: 폴더별 8회 평균 통계"""
    headers = [
        "명령어", "분석 파일 수", "유효 파일 수",
        "T0 평균 (s)", "T1 평균 (s)", "T2 평균 (s)", "T3 평균 (s)",
        "E2E 평균 (s)", "E2E 최소 (s)", "E2E 최대 (s)", "오류 파일"
    ]
    col_widths = [28, 14, 14, 14, 14, 14, 14, 16, 16, 16, 30]

    _write_header_row(ws, 1, headers, HEADER_FILL, HEADER_FONT)
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    row = 2
    for command, stats in folder_stats.items():
        values = [
            command,
            stats.total_files,
            stats.valid_count,
            stats.avg_T0,
            stats.avg_T1,
            stats.avg_T2,
            stats.avg_T3,
            stats.avg_E2E,
            stats.min_E2E,
            stats.max_E2E,
            ", ".join(stats.error_files) if stats.error_files else "-",
        ]
        _write_data_row(ws, row, values)
        row += 1

    # E2E 평균 컬럼 조건부 서식
    if row > 2:
        ws.conditional_formatting.add(
            f"H2:H{row}",
            ColorScaleRule(
                start_type="min", start_color="63BE7B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="F8696B",
            )
        )

    ws.freeze_panes = "A2"


def _write_header_row(ws, row: int, headers: list, fill, font):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER
        cell.border = BORDER_THIN


def _write_data_row(ws, row: int, values: list, fill=None, font=None):
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.alignment = CENTER if col > 1 else LEFT
        cell.border = BORDER_THIN
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        # 숫자 포맷
        if isinstance(v, float):
            cell.number_format = "0.000"
